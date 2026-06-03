"""Tests for uiao.adapters.zta.render — all five output formats."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from uiao.adapters.zta.ingest import load_report
from uiao.adapters.zta.render import (
    render_executive,
    render_html,
    render_markdown,
    resolve_formats,
    write_outputs,
)
from uiao.adapters.zta.render.csv_out import worklist_rows
from uiao.adapters.zta.triage import build_triage

FIXTURES = Path(__file__).parent / "fixtures" / "zta"
MINI_JSON = FIXTURES / "sample-mini.json"


@pytest.fixture()
def triage():
    return build_triage(load_report(MINI_JSON))


def test_resolve_formats_all_and_validation():
    assert resolve_formats(["all"]) == ["exec", "md", "xlsx", "html", "csv"]
    assert resolve_formats(None) == ["exec", "md", "xlsx", "html", "csv"]
    assert resolve_formats(["csv", "md"]) == ["md", "csv"]
    with pytest.raises(ValueError):
        resolve_formats(["bogus"])


def test_executive_mentions_posture_and_caveat(triage):
    md = render_executive(triage)
    assert "Executive Summary" in md
    assert "3 actionable findings" in md
    assert "premium/P2-SKU-gated" in md


def test_markdown_digest_has_sections_and_worklist(triage):
    md = render_markdown(triage)
    assert "## Outcome" in md
    assert "## Pillar × Status" in md
    assert "## Applicability" in md
    assert "Weak authentication methods are disabled" in md
    # Summary block must not leak into the digest.
    assert "999" not in md


def test_html_is_self_contained_and_escaped(triage):
    html = render_html(triage)
    assert html.startswith("<!doctype html>")
    assert "<style>" in html  # inline CSS, no external assets
    assert "Fixture Corp" in html
    assert "premium" in html


def test_csv_rows_cover_worklist(triage):
    rows = worklist_rows(triage)
    assert len(rows) == 3
    assert {r["TestId"] for r in rows} == {
        "21001",
        "20606e75-05c4-48c0-9d97-add6daa2109a",
        "24006",
    }
    assert all("Applicability" in r for r in rows)


def test_write_outputs_creates_all_files(tmp_path: Path, triage):
    paths = write_outputs(triage, tmp_path)
    names = sorted(p.name for p in paths)
    assert len(names) == 5
    for p in paths:
        assert p.exists() and p.stat().st_size > 0


def test_xlsx_workbook_structure(tmp_path: Path, triage):
    paths = write_outputs(triage, tmp_path, ["xlsx"])
    wb = openpyxl.load_workbook(paths[0])
    assert wb.sheetnames == ["Summary", "Worklist", "All Tests", "By Pillar"]
    worklist = wb["Worklist"]
    assert worklist.max_row == 4  # header + 3 findings
    all_tests = wb["All Tests"]
    assert all_tests.max_row == 8  # header + 7 tests
    assert worklist.freeze_panes == "A2"
