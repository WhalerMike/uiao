"""Tests for uiao.adapters.zta.render.dashboard — charts, workbook, Power BI CSV."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from uiao.adapters.zta.analytics import FACT_COLUMNS, build_facts, findings
from uiao.adapters.zta.ingest import load_report
from uiao.adapters.zta.render.dashboard import write_dashboard

FIXTURES = Path(__file__).parent / "fixtures" / "zta"
MINI_JSON = FIXTURES / "sample-mini.json"


@pytest.fixture()
def report():
    return load_report(MINI_JSON)


def test_write_dashboard_produces_all_outputs(tmp_path: Path, report):
    out = write_dashboard(report, tmp_path)
    assert out.workbook.exists() and out.workbook.stat().st_size > 0
    assert out.powerbi_csv.exists() and out.powerbi_csv.stat().st_size > 0
    assert len(out.charts) == 4
    for chart in out.charts:
        assert chart.exists() and chart.suffix == ".png" and chart.stat().st_size > 0


def test_workbook_sheets_and_embedded_charts(tmp_path: Path, report):
    out = write_dashboard(report, tmp_path)
    wb = openpyxl.load_workbook(out.workbook)
    assert wb.sheetnames == ["Dashboard", "Findings", "By Pillar", "All Tests", "Methodology"]
    assert len(wb["Dashboard"]._images) == 4
    # Findings sheet holds exactly the findings (+ header).
    assert wb["Findings"].max_row == len(findings(build_facts(report))) + 1
    # All Tests holds every test (+ header).
    assert wb["All Tests"].max_row == len(report.tests) + 1


def test_by_pillar_findings_total_reconciles(tmp_path: Path, report):
    out = write_dashboard(report, tmp_path)
    wb = openpyxl.load_workbook(out.workbook)
    bp = wb["By Pillar"]
    header = [c.value for c in bp[1]]
    fcol = header.index("Findings") + 1
    total = sum(bp.cell(row=r, column=fcol).value for r in range(2, bp.max_row + 1))
    assert total == len(findings(build_facts(report)))


def test_powerbi_csv_is_tidy(tmp_path: Path, report):
    out = write_dashboard(report, tmp_path)
    with out.powerbi_csv.open(encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == len(report.tests)  # one row per test
    for col in (*FACT_COLUMNS, "Tenant", "ExecutedAt", "IsDemo"):
        assert col in rows[0]
