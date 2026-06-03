"""Dashboard renderer — charts/heatmap, enriched workbook, and Power BI CSV.

Produces, from a parsed report:

* PNG charts (status, findings-by-pillar, pillar×risk heatmap, severity×effort
  remediation quadrant) — white background, house palette;
* an Excel workbook with the charts embedded on a Dashboard sheet plus Findings
  (ranked, severity color-scale, priority-bucket fills), By Pillar, All Tests,
  and a Methodology sheet;
* a tidy one-row-per-test **Power BI-ready CSV**.

matplotlib/numpy are stub-less (see the pyproject mypy override); openpyxl ships
types via ``types-openpyxl``.
"""

from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
from matplotlib.axes import Axes  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.drawing.image import Image as XLImage  # noqa: E402
from openpyxl.formatting.rule import ColorScaleRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.worksheet import Worksheet  # noqa: E402

from uiao.adapters.zta.analytics import FACT_COLUMNS, Fact, build_facts, fact_record, findings  # noqa: E402
from uiao.adapters.zta.model import STATUS_ORDER, ZtReport  # noqa: E402
from uiao.adapters.zta.render._common import sanitize_basename  # noqa: E402
from uiao.adapters.zta.triage import RISK_ORDER  # noqa: E402

# Palette: white background; navy/teal/amber/red house style.
C_NAVY, C_AMBER, C_RED, C_GREEN, C_GREY, C_BLUE = "#1F3864", "#BF8F00", "#C00000", "#548235", "#A6A6A6", "#2E75B6"
RISK_COLOR: dict[str, str] = {"High": C_RED, "Medium": C_AMBER, "Low": C_GREEN}
STATUS_COLOR: dict[str, str] = {
    "Failed": C_RED,
    "Investigate": C_AMBER,
    "Passed": C_GREEN,
    "Skipped": C_GREY,
    "Planned": C_BLUE,
    "Error": C_NAVY,
}


@dataclass(frozen=True)
class DashboardOutputs:
    """Paths written by :func:`write_dashboard`."""

    workbook: Path
    powerbi_csv: Path
    charts: list[Path]


# --------------------------------------------------------------------------- #
# Charts
# --------------------------------------------------------------------------- #
def _style_ax(ax: Axes, title: str) -> None:
    ax.set_title(title, color=C_NAVY, fontsize=12, fontweight="bold")
    ax.set_facecolor("white")
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)


def _pillars(facts: list[Fact]) -> list[str]:
    return sorted({f.pillar for f in facts})


def chart_status(report: ZtReport, path: Path) -> Path:
    counts = Counter(t.test_status for t in report.tests)
    labels = [s for s in STATUS_ORDER if counts.get(s)]
    values = [counts[s] for s in labels]
    fig, ax = plt.subplots(figsize=(6, 3.4), facecolor="white")
    bars = ax.bar(labels, values, color=[STATUS_COLOR.get(s, C_GREY) for s in labels])
    ax.bar_label(bars, padding=2, color=C_NAVY, fontsize=9)
    _style_ax(ax, "Outcomes (recomputed from per-test results)")
    ax.set_ylabel("checks")
    fig.tight_layout()
    fig.savefig(path, dpi=130, facecolor="white")
    plt.close(fig)
    return path


def chart_by_pillar(facts: list[Fact], path: Path) -> Path:
    pillars = _pillars(facts)
    idx = {p: i for i, p in enumerate(pillars)}
    data: dict[str, list[float]] = {risk: [0.0] * len(pillars) for risk in RISK_ORDER}
    for f in facts:
        if f.is_finding and f.risk in data:
            data[f.risk][idx[f.pillar]] += 1
    fig, ax = plt.subplots(figsize=(6.6, 3.6), facecolor="white")
    bottom = np.zeros(len(pillars))
    for risk in RISK_ORDER:
        vals = np.array(data[risk])
        ax.bar(pillars, vals, bottom=bottom, label=risk, color=RISK_COLOR[risk])
        bottom += vals
    _style_ax(ax, "Findings by pillar (stacked by risk)")
    ax.set_ylabel("findings")
    ax.legend(frameon=False, fontsize=8)
    plt.setp(ax.get_xticklabels(), rotation=20, ha="right")
    fig.tight_layout()
    fig.savefig(path, dpi=130, facecolor="white")
    plt.close(fig)
    return path


def chart_heatmap(facts: list[Fact], path: Path) -> Path:
    pillars = _pillars(facts)
    risks = list(RISK_ORDER)
    pidx = {p: i for i, p in enumerate(pillars)}
    ridx = {r: i for i, r in enumerate(risks)}
    grid = np.zeros((len(pillars), len(risks)), dtype=int)
    for f in facts:
        if f.is_finding and f.risk in ridx:
            grid[pidx[f.pillar], ridx[f.risk]] += 1
    fig, ax = plt.subplots(figsize=(5.2, 3.8), facecolor="white")
    im = ax.imshow(grid, cmap="Reds", aspect="auto")
    ax.set_xticks(range(len(risks)), labels=risks)
    ax.set_yticks(range(len(pillars)), labels=pillars)
    vmax = int(grid.max()) if grid.size else 0
    for i in range(len(pillars)):
        for j in range(len(risks)):
            val = int(grid[i, j])
            color = "white" if vmax and val > vmax * 0.6 else C_NAVY
            ax.text(j, i, str(val), ha="center", va="center", color=color, fontsize=9)
    _style_ax(ax, "Finding-count heatmap (pillar × risk)")
    fig.colorbar(im, ax=ax, shrink=0.8, label="findings")
    fig.tight_layout()
    fig.savefig(path, dpi=130, facecolor="white")
    plt.close(fig)
    return path


def chart_quadrant(facts: list[Fact], path: Path) -> Path:
    rng = np.random.default_rng(7)  # deterministic jitter
    fig, ax = plt.subplots(figsize=(6.0, 3.8), facecolor="white")
    for f in findings(facts):
        ax.scatter(
            f.effort_weight + rng.uniform(-0.12, 0.12),
            f.severity,
            c=RISK_COLOR.get(f.risk, C_GREY),
            alpha=0.75,
            edgecolors="white",
            linewidths=0.5,
            s=42,
        )
    ax.axhline(60, color=C_GREY, linewidth=0.8, linestyle="--")
    ax.axvline(2, color=C_GREY, linewidth=0.8, linestyle="--")
    ax.set_xticks([1, 2, 3], labels=["Low", "Medium", "High"])
    ax.set_xlabel("remediation effort")
    ax.set_ylabel("severity")
    ax.text(0.95, 102, "quick wins", color=C_GREEN, fontsize=8, fontweight="bold")
    ax.text(2.55, 102, "major projects", color=C_RED, fontsize=8, fontweight="bold")
    _style_ax(ax, "Remediation quadrant (severity × effort)")
    fig.tight_layout()
    fig.savefig(path, dpi=130, facecolor="white")
    plt.close(fig)
    return path


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=C_NAVY.lstrip("#"))
_WRAP = Alignment(vertical="top", wrap_text=True)
_BUCKET_FILL: dict[str, PatternFill] = {
    "P1": PatternFill("solid", fgColor="F4CCCC"),
    "P2": PatternFill("solid", fgColor="FCE5CD"),
    "P3": PatternFill("solid", fgColor="FFF2CC"),
    "P4": PatternFill("solid", fgColor="D9EAD3"),
}
_COL_WIDTH: dict[str, int] = {
    "Title": 58,
    "MinimumLicense": 26,
    "ApplicabilityNote": 46,
    "Category": 22,
    "SfiPillar": 24,
}


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _write_table(ws: Worksheet, facts: list[Fact]) -> None:
    columns = list(FACT_COLUMNS)
    for c, name in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=name)
        ws.column_dimensions[get_column_letter(c)].width = _COL_WIDTH.get(name, 13)
    for r, fact in enumerate(facts, start=2):
        record = fact_record(fact)
        for c, name in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=record[name])
            if name == "Title":
                cell.alignment = _WRAP
            elif name == "PriorityBucket":
                fill = _BUCKET_FILL.get(str(record[name]))
                if fill is not None:
                    cell.fill = fill
    _style_header(ws, len(columns))
    if facts:
        col = get_column_letter(columns.index("SeverityScore") + 1)
        ws.conditional_formatting.add(
            f"{col}2:{col}{len(facts) + 1}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="D9EAD3",
                mid_type="num",
                mid_value=50,
                mid_color="FFF2CC",
                end_type="num",
                end_value=100,
                end_color="F4CCCC",
            ),
        )


def _write_by_pillar(ws: Worksheet, facts: list[Fact]) -> None:
    status_by_pillar: dict[str, Counter[str]] = defaultdict(Counter)
    sev_by_pillar: dict[str, list[int]] = defaultdict(list)
    for f in facts:
        status_by_pillar[f.pillar][f.status] += 1
        if f.is_finding:
            sev_by_pillar[f.pillar].append(f.severity)
    columns = ["Pillar", *STATUS_ORDER, "Findings", "AvgSeverity", "MaxSeverity"]
    for c, name in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=name)
    for r, pillar in enumerate(sorted(status_by_pillar), start=2):
        counts = status_by_pillar[pillar]
        sevs = sev_by_pillar.get(pillar, [])
        values: list[str | int | float] = [
            pillar,
            *[counts.get(s, 0) for s in STATUS_ORDER],
            len(sevs),
            round(sum(sevs) / len(sevs), 1) if sevs else 0,
            max(sevs) if sevs else 0,
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=value)
    ws.column_dimensions["A"].width = 16
    _style_header(ws, len(columns))


def _write_methodology(ws: Worksheet, report: ZtReport) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 92
    lines: list[tuple[str, str]] = [
        ("Source", "Microsoft Zero Trust Assessment (github.com/microsoft/zerotrustassessment)"),
        ("Tenant", f"{report.tenant_name or ''} ({report.domain or ''})"),
        ("Tool version", report.current_version or ""),
        ("Executed at", report.executed_at or ""),
        ("Demo data", "yes" if report.is_demo else "no"),
        ("", ""),
        ("Outcome handling", "Roll-ups recomputed from Tests[]; the report's TestResultSummary is not trusted."),
        ("Findings", "Only Failed and Investigate are findings. Skipped/Planned/Error are non-gaps."),
        (
            "Severity (0-100)",
            "risk_weight x status_weight x 10. Risk: High=10 Medium=6 Low=3. "
            "Status: Failed=1.0 Investigate=0.6; all others (incl. Error)=0.",
        ),
        ("Effort weight", "TestImplementationCost: Low=1 Medium=2 High=3 (blank -> 2)."),
        ("Priority score", "severity / effort_weight. Higher = remediate sooner (quick wins)."),
        ("Priority buckets", "P1>=50, P2>=30, P3>=15, P4>0."),
        (
            "Applicability",
            "Derived from SkippedReason + premium/P2 SKU gating. Premium findings are "
            "flagged: verify GCC/FedRAMP-Moderate availability before treating as a real gap.",
        ),
        (
            "Boundary",
            "Findings are Controlled tenant data — keep in-boundary, do not publish raw results. "
            "Remediation is proposed, not applied (federal L3 actuation ceiling).",
        ),
    ]
    for r, (key, value) in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=key).font = Font(bold=bool(key))
        ws.cell(row=r, column=2, value=value).alignment = _WRAP


def write_workbook(report: ZtReport, facts: list[Fact], charts: list[Path], path: Path) -> Path:
    """Write the enriched dashboard workbook to ``path``."""
    wb = Workbook()
    dash = wb.active
    assert dash is not None
    dash.title = "Dashboard"
    dash.sheet_view.showGridLines = False
    dash.cell(row=1, column=1, value=f"Zero Trust Assessment — {report.tenant_name or ''}").font = Font(
        bold=True, size=14, color=C_NAVY.lstrip("#")
    )
    for chart_path, anchor in zip(charts, ("A3", "J3", "A23", "J23"), strict=False):
        dash.add_image(XLImage(str(chart_path)), anchor)
    _write_table(wb.create_sheet("Findings"), findings(facts))
    _write_by_pillar(wb.create_sheet("By Pillar"), facts)
    _write_table(wb.create_sheet("All Tests"), facts)
    _write_methodology(wb.create_sheet("Methodology"), report)
    wb.save(str(path))
    return path


def write_powerbi_csv(report: ZtReport, facts: list[Fact], path: Path) -> Path:
    """Write a tidy one-row-per-test fact table for Power BI import."""
    context = {
        "Tenant": report.tenant_name or "",
        "Domain": report.domain or "",
        "ToolVersion": report.current_version or "",
        "ExecutedAt": report.executed_at or "",
        "IsDemo": "yes" if report.is_demo else "no",
    }
    columns = [*context.keys(), *FACT_COLUMNS]
    with path.open("w", newline="", encoding="utf-8-sig") as fh:  # BOM: Power BI auto-detects UTF-8
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        for fact in facts:
            writer.writerow({**context, **fact_record(fact)})
    return path


def write_dashboard(report: ZtReport, out_dir: Path) -> DashboardOutputs:
    """Render the full dashboard (charts + workbook + Power BI CSV) into ``out_dir``."""
    facts = build_facts(report)
    out_dir.mkdir(parents=True, exist_ok=True)
    charts_dir = out_dir / "charts"
    charts_dir.mkdir(parents=True, exist_ok=True)
    base = sanitize_basename(report)
    charts = [
        chart_status(report, charts_dir / f"{base}-status.png"),
        chart_by_pillar(facts, charts_dir / f"{base}-by-pillar.png"),
        chart_heatmap(facts, charts_dir / f"{base}-heatmap.png"),
        chart_quadrant(facts, charts_dir / f"{base}-quadrant.png"),
    ]
    return DashboardOutputs(
        workbook=write_workbook(report, facts, charts, out_dir / f"{base}-dashboard.xlsx"),
        powerbi_csv=write_powerbi_csv(report, facts, out_dir / f"{base}-powerbi.csv"),
        charts=charts,
    )
