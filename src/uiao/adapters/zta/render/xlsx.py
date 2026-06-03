"""Excel workbook renderer (openpyxl).

Produces a triage workbook with four sheets: Summary, Worklist, All Tests,
and By Pillar. Header rows are bold + frozen with an autofilter so a team can
sort and slice the findings.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from uiao.adapters.zta.applicability import derive_applicability, is_premium_gated, license_str
from uiao.adapters.zta.model import STATUS_ORDER, ZtTest
from uiao.adapters.zta.triage import Triage, id_namespace

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")  # navy, matches repo palette
_WRAP = Alignment(vertical="top", wrap_text=True)

_TEST_COLUMNS: list[tuple[str, int]] = [
    ("TestId", 14),
    ("IdNamespace", 12),
    ("Pillar", 14),
    ("SfiPillar", 24),
    ("Category", 22),
    ("Title", 60),
    ("Risk", 8),
    ("Status", 12),
    ("Impact", 9),
    ("Effort", 9),
    ("MinimumLicense", 28),
    ("Premium", 9),
    ("Applicability", 18),
]


def _test_row(t: ZtTest) -> list[str]:
    return [
        t.test_id,
        id_namespace(t.test_id),
        t.pillar,
        t.test_sfi_pillar or "",
        t.test_category or "",
        t.test_title,
        t.test_risk,
        t.test_status,
        t.test_impact,
        t.test_implementation_cost,
        license_str(t),
        "yes" if is_premium_gated(t) else "no",
        derive_applicability(t).value,
    ]


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _write_tests_sheet(ws: Worksheet, tests: list[ZtTest]) -> None:
    for col, (name, width) in enumerate(_TEST_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
        ws.column_dimensions[get_column_letter(col)].width = width
    for r, t in enumerate(tests, start=2):
        for c, value in enumerate(_test_row(t), start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if _TEST_COLUMNS[c - 1][0] == "Title":
                cell.alignment = _WRAP
    _style_header(ws, len(_TEST_COLUMNS))


def _write_summary_sheet(ws: Worksheet, triage: Triage) -> None:
    r = triage.report
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60
    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    facts: list[tuple[str, str]] = [
        ("Tenant", r.tenant_name or ""),
        ("Domain", r.domain or ""),
        ("Tool version", r.current_version or ""),
        ("Executed at", r.executed_at or ""),
        ("Demo data", "yes" if r.is_demo else "no"),
        ("Total checks", str(len(r.tests))),
        ("Actionable (Failed/Investigate)", str(triage.actionable_total)),
        ("Skipped (non-gap)", str(triage.skipped_total)),
        ("Planned (non-gap)", str(triage.planned_total)),
        ("Premium/P2-SKU-gated", str(triage.premium_gated_total)),
        (
            "TestId: numeric / GUID",
            f"{triage.namespace_counts.get('numeric', 0)} / {triage.namespace_counts.get('guid', 0)}",
        ),
    ]
    row = 2
    for key, value in facts:
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for status in STATUS_ORDER:
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=triage.status_counts.get(status, 0))
        row += 1
    _style_header(ws, 2)


def _write_pillar_sheet(ws: Worksheet, triage: Triage) -> None:
    ws.cell(row=1, column=1, value="Pillar")
    ws.column_dimensions["A"].width = 16
    for c, status in enumerate(STATUS_ORDER, start=2):
        ws.cell(row=1, column=c, value=status)
        ws.column_dimensions[get_column_letter(c)].width = 12
    for r, (pillar, counts) in enumerate(triage.pillar_crosstab.items(), start=2):
        ws.cell(row=r, column=1, value=pillar)
        for c, status in enumerate(STATUS_ORDER, start=2):
            ws.cell(row=r, column=c, value=counts.get(status, 0))
    _style_header(ws, len(STATUS_ORDER) + 1)


def write_xlsx(triage: Triage, path: Path) -> Path:
    """Write the triage workbook to ``path``; return the path."""
    wb = Workbook()
    summary = wb.active
    assert summary is not None
    summary.title = "Summary"
    _write_summary_sheet(summary, triage)

    _write_tests_sheet(wb.create_sheet("Worklist"), triage.worklist)
    _write_tests_sheet(wb.create_sheet("All Tests"), triage.report.tests)
    _write_pillar_sheet(wb.create_sheet("By Pillar"), triage)

    wb.save(str(path))
    return path
